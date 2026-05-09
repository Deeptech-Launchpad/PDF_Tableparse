# =============================================================================
# core/exporter.py
# Handles exporting results to individual JSON files, a combined JSON file,
# and a styled Excel workbook.
# =============================================================================

import json
import os
import re
from typing import List, Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.constants import EXCEL_FILENAME, EXCEL_SHEET_NAME, COMBINED_JSON_FILENAME
from utils.logger import logger

def natural_sort_key(s):
    """Utility for sorting strings with numbers numerically (Json-1, Json-2, Json-10)."""
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', s)]


# ---------------------------------------------------------------------------
# JSON Export
# ---------------------------------------------------------------------------

def save_individual_json(record: Dict[str, Any], json_path: str) -> None:
    """
    Save a single extracted record as an individual JSON file.

    Args:
        record: Dictionary containing id, pdf_name, page, image_path, text, etc.
        json_path: Full file path where the JSON should be saved.
    """
    try:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(record, f, indent=4, ensure_ascii=False)
        logger.debug(f"Saved JSON: {os.path.basename(json_path)}")
    except Exception as exc:
        logger.error(f"Failed to save JSON {json_path}: {exc}")


def save_combined_json(records: List[Dict[str, Any]], output_dir: str) -> str:
    """
    Save all records into a single combined JSON file.

    Args:
        records: List of all result dictionaries.
        output_dir: Directory where combined JSON will be saved.

    Returns:
        Path to the combined JSON file.
    """
    combined_path = os.path.join(output_dir, COMBINED_JSON_FILENAME)
    try:
        with open(combined_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=4, ensure_ascii=False)
        logger.info(f"Combined JSON saved: {combined_path} ({len(records)} records)")
    except Exception as exc:
        logger.error(f"Failed to save combined JSON: {exc}")
    return combined_path


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def save_excel(records: List[Dict[str, Any]], excel_path: str) -> str:
    """
    Export all records to a formatted .xlsx Excel file.
    If a record contains a structured table (headers/rows), it expands it 
    into a beautiful, printable Excel table instead of a JSON string.
    """
    if not records:
        return excel_path

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME
    
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    
    # Styling Constants
    header_fill = PatternFill("solid", fgColor="D32F2F") # Red Section
    column_fill = PatternFill("solid", fgColor="F5F5DC") # Beige Header
    row_fill_a = PatternFill("solid", fgColor="FFFFFF")
    row_fill_b = PatternFill("solid", fgColor="FAFAFA")
    
    border_side = Side(style="thin", color="D1D5DB")
    table_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    current_row = 1
    
    for rec in records:
        # Check if this record is a structured table
        # We look for the presence of 'headers' and 'rows' in the fields
        is_table = False
        table_data = None
        
        # In our system, structured tables are stored directly in 'fields'
        if isinstance(rec, dict) and "headers" in rec and "rows" in rec:
            is_table = True
            table_data = rec
        
        if is_table:
            # 1. Write Section Title
            cell = ws.cell(row=current_row, column=1, value=table_data.get("section", "Data Table"))
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(table_data["headers"]))
            current_row += 1
            
            # 2. Write Accuracy Note if exists
            if table_data.get("accuracy_note") and table_data["accuracy_note"] != "N/A":
                cell = ws.cell(row=current_row, column=1, value=table_data["accuracy_note"])
                cell.font = Font(italic=True, size=10)
                cell.alignment = Alignment(horizontal="left")
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(table_data["headers"]))
                current_row += 1

            # 3. Write Headers
            for col_idx, h in enumerate(table_data["headers"], start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=h)
                cell.font = Font(bold=True, size=10)
                cell.fill = column_fill
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            current_row += 1
            
            # 4. Write Rows
            for r_idx, row in enumerate(table_data["rows"]):
                fill = row_fill_a if r_idx % 2 == 0 else row_fill_b
                for c_idx, val in enumerate(row, start=1):
                    # Handle cases where row might be shorter than headers
                    if c_idx > len(table_data["headers"]): break
                    
                    cleaned_val = ILLEGAL_CHARACTERS_RE.sub('', str(val))
                    cell = ws.cell(row=current_row, column=c_idx, value=cleaned_val)
                    cell.border = table_border
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                current_row += 1
                
            # Add some spacing between tables
            current_row += 2
        else:
            # Fallback for standard key-value records (like product metadata)
            for k, v in rec.items():
                ws.cell(row=current_row, column=1, value=k).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=str(v))
                current_row += 1
            current_row += 1

    wb.save(excel_path)
    logger.info(f"PDF-Ready Excel saved: {excel_path}")
    return excel_path


def _apply_excel_styling(ws, excel_columns) -> None:
    """Apply header, alternating rows, borders, and freeze to an openpyxl worksheet."""
    header_fill   = PatternFill("solid", fgColor="1A1F36")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin", color="2E3250")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Apply header row
    for col_idx, (label, key, width) in enumerate(excel_columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value       = label
        cell.fill        = header_fill
        cell.font        = header_font
        cell.alignment   = center_align
        cell.border      = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # Alternating row fill
    row_fill_a = PatternFill("solid", fgColor="12152B")
    row_fill_b = PatternFill("solid", fgColor="1C2040")
    data_font  = Font(color="D0D4F0", size=10)

    for row_idx in range(2, ws.max_row + 1):
        fill = row_fill_a if row_idx % 2 == 0 else row_fill_b
        for col_idx in range(1, len(excel_columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = fill
            cell.font      = data_font
            cell.border    = border
            cell.alignment = left_align
        ws.row_dimensions[row_idx].height = 40

    ws.freeze_panes = "A2"
